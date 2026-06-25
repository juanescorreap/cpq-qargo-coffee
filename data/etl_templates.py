"""
ETL: data/files-to-explore/ + data/raw/ → data/templates/
Implementa el plan documentado en data/ETL_TEMPLATES_PLAN.md.

Fuentes:
  data/files-to-explore/  → fabricantes, distribuidores, precios de ruta
  data/raw/ingredients.xlsx → 165 ingredientes con precios canónicos (fuente primaria)
  data/raw/conversions.xlsx → 79 conversiones de unidad de receta

Uso:
    python data/etl_templates.py           # todas las fases
    python data/etl_templates.py --phase 0 # solo crosswalk
    python data/etl_templates.py --phase 1 # regions, manufacturers, distributors
    python data/etl_templates.py --phase 2 # stores_regions, ingredient_prices, supply_routes
    python data/etl_templates.py --phase 3 # supplier_refs, unit_conversions, route_prices
    python data/etl_templates.py --phase 4 # supply_route_assignments
"""

import argparse
import csv
import difflib
import re
import statistics
from datetime import date, datetime
from pathlib import Path

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).parent
SRC        = ROOT / "_archive" / "source_documents"
RAW        = ROOT / "raw"
TMPL       = ROOT / "templates"
CANONICAL  = ROOT / "ingredients-v1.xlsx"
CROSSWALK  = ROOT / "crosswalk.csv"

BRIDOR         = SRC / "2025_Q4 - Bridor Rebate Report.xlsx"
BINDI          = SRC / "Comparison Bindi - Report (6_19_2026).xlsx"
PREGEL         = SRC / "PreGel Consolidate 2026.xlsx"
WAFFLES        = SRC / "Goolden Waffles Consolidate 2026.xlsx"
STORES_XLS     = SRC / "lista de tiendas.xlsx"
RAW_INGREDIENTS = RAW / "ingredients.xlsx"
RAW_CONVERSIONS = RAW / "conversions.xlsx"

TMPL.mkdir(exist_ok=True)

# ── Normalisation tables ──────────────────────────────────────────────────

MANUFACTURER_MAP: dict[str, dict] = {
    "bridor":        {"name": "Bridor",         "country_code": "FR"},
    "bindi":         {"name": "Bindi",           "country_code": "IT"},
    "coffee":        {"name": "Kimbo",           "country_code": "IT"},
    "pregel":        {"name": "PreGel",          "country_code": "IT"},
    "golden waffles":{"name": "Golden Waffles",  "country_code": "US"},
    "goolden waffles":{"name": "Golden Waffles", "country_code": "US"},
}

# Group values in Bindi that are DISTRIBUTORS (not manufacturers)
BINDI_DISTRIBUTORS: set[str] = {
    "food related", "french bakery", "greco", "greco & sons",
    "pointe dairy", "tcw", "igf", "primizie",
}

DISTRIBUTOR_NORM: dict[str, str] = {
    "greco & sons":  "Greco & Sons",
    "greco":         "Greco & Sons",
    "food related":  "Food Related",
    "french bakery": "French Bakery",
    "pointe dairy":  "Pointe Dairy",
    "tcw":           "TCW",
    "igf":           "IGF",
    "primizie":      "Primizie",
}

STATE_TO_REGION: dict[str, str] = {
    "CA":   "California",
    "FL":   "Florida",
    "TX":   "Texas",
    "MI":   "Michigan",
    "OH":   "Ohio",
    "IL":   "Illinois",
    "NV":   "Nevada",
    "D.C.": "Washington D.C.",
    "DC":   "Washington D.C.",
}

STORE_NAME_TO_CODE: dict[str, str] = {
    "long beach":           "2-LB-CA",
    "fountain valley":      "1-FV-CA",
    "tampa":                "3-TM-FL",
    "farragut":             "4-WC-DC",
    "washington":           "4-WC-DC",
    "lafayette":            "4-WC-DC",
    "berkeley":             "5-BK-CA",
    "detroit":              "6-DT-MI",
    "edinburg":             "7-ED-TX",
    "edimburg":             "7-ED-TX",
    "westerville":          "8-WV-OH",
    "meijer 215":           "10-BL-IL",
    "meijer boolingbrook boughton 215": "10-BL-IL",
    "san antonio":          "11-SA-TX",
    "dearborn":             "12-DB-MI",
    "meijer 169":           "13-BL-IL",
    "meijer bolingbrook weber 169": "13-BL-IL",
    "meijer 182":           "14-SC-IL",
    "meijer saint charles 182": "14-SC-IL",
    "orland park":          "15-OP-IL",
    "grand prairie":        "16-GP-TX",
    "gran praire":          "16-GP-TX",
    "vegas":                "17-VG-NV",
    "canton":               "18-CN-MI",
}

# Stores that existed in reports but are NOT in the current 17-store network
SKIP_STORES: set[str] = {
    "santa monica", "san jose", "denver", "west palm beach",
    "cooper city", "st pete", "st. pete", "st. petersburg",
    "st petersburg", "meijer 183", "dearborn 22022",
}

# ── Helpers ────────────────────────────────────────────────────────────────

def clean_price(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(float(val), 4) if float(val) > 0 else None
    s = str(val).strip().replace("$", "").replace(",", ".")
    try:
        f = float(s)
        return round(f, 4) if f > 0 else None
    except ValueError:
        return None


def clean_code(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
    return str(val).strip() or None


def clean_external_name(raw: str) -> str:
    """Strip supplier SKU prefix: '0026P - CHEESECAKE ALLE FRAGOLE ' → 'Cheesecake Alle Fragole'"""
    s = re.sub(r'^[\w\-\.]+\s*[-–]\s*', '', str(raw)).strip()
    return s.title() if s else str(raw).strip().title()


def norm_state(raw: str) -> str:
    return "DC" if raw in ("D.C.", "DC") else raw.strip().upper()


def norm_store_name(raw: str) -> str | None:
    key = str(raw).strip().lower()
    # Direct match
    if key in STORE_NAME_TO_CODE:
        return STORE_NAME_TO_CODE[key]
    # Partial match
    for k, v in STORE_NAME_TO_CODE.items():
        if k in key or key in k:
            return v
    return None


def clean_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def load_wb_rows(path: Path, sheet: str) -> tuple[list, list[tuple]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c else "" for c in rows[0]]
    data = [r for r in rows[1:] if any(v is not None for v in r)]
    return headers, data


def write_csv(path: Path, headers: list[str], rows: list[dict]) -> int:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow({h: row.get(h, "") for h in headers})
    return len(rows)


def today_iso() -> str:
    return date.today().isoformat()


# ══════════════════════════════════════════════════════════════════════════
# FASE 0 — Crosswalk: external names → canonical ingredient names
# ══════════════════════════════════════════════════════════════════════════

def build_crosswalk() -> dict[str, str]:
    """Returns {external_name_lower: canonical_name}"""
    # Load canonical names
    wb = openpyxl.load_workbook(CANONICAL, data_only=True)
    ws = wb.active
    h = [c.value for c in ws[1]]
    ni = h.index("name")
    canonicals = [
        str(r[ni]).strip()
        for r in ws.iter_rows(min_row=2, values_only=True)
        if r[ni]
    ]
    canonical_lower = [c.lower() for c in canonicals]

    # Collect all external names from source files
    externals: set[str] = set()

    # Bindi Products Database
    try:
        h_b, rows_b = load_wb_rows(BINDI, "Products Database")
        item_i = next(i for i, x in enumerate(h_b) if "item" in x.lower())
        for r in rows_b:
            if r[item_i]:
                externals.add(clean_external_name(str(r[item_i])))
    except Exception:
        pass

    # Bindi Revised
    try:
        h_r, rows_r = load_wb_rows(BINDI, "Revised Prod Database")
        item_i = next(i for i, x in enumerate(h_r) if "item" in x.lower())
        for r in rows_r:
            if r[item_i]:
                externals.add(clean_external_name(str(r[item_i])))
    except Exception:
        pass

    # PreGel
    try:
        h_p, rows_p = load_wb_rows(PREGEL, "Products Database")
        item_i = next(i for i, x in enumerate(h_p) if "item" in x.lower())
        for r in rows_p:
            if r[item_i]:
                externals.add(clean_external_name(str(r[item_i])))
    except Exception:
        pass

    # Golden Waffles
    try:
        h_g, rows_g = load_wb_rows(WAFFLES, "Products Database")
        item_i = next(i for i, x in enumerate(h_g) if "item" in x.lower())
        for r in rows_g:
            if r[item_i]:
                externals.add(clean_external_name(str(r[item_i])))
    except Exception:
        pass

    # Bridor
    try:
        h_br, rows_br = load_wb_rows(BRIDOR, "Data")
        desc_i = next(i for i, x in enumerate(h_br) if "description" in x.lower())
        for r in rows_br:
            if r[desc_i]:
                externals.add(clean_external_name(str(r[desc_i])))
    except Exception:
        pass

    # Build crosswalk with fuzzy matching
    mapping: dict[str, str] = {}  # external_lower → canonical
    crosswalk_rows = []

    for ext in sorted(externals):
        ext_lower = ext.lower()
        # Exact match
        if ext_lower in canonical_lower:
            idx = canonical_lower.index(ext_lower)
            mapping[ext_lower] = canonicals[idx]
            crosswalk_rows.append({
                "external_name": ext,
                "canonical_name": canonicals[idx],
                "match_score": 1.0,
                "status": "exact",
            })
            continue
        # Fuzzy match
        matches = difflib.get_close_matches(ext_lower, canonical_lower, n=1, cutoff=0.55)
        if matches:
            idx = canonical_lower.index(matches[0])
            score = difflib.SequenceMatcher(None, ext_lower, matches[0]).ratio()
            status = "auto" if score >= 0.72 else "review"
            if score >= 0.72:
                mapping[ext_lower] = canonicals[idx]
            crosswalk_rows.append({
                "external_name": ext,
                "canonical_name": canonicals[idx],
                "match_score": round(score, 3),
                "status": status,
            })
        else:
            crosswalk_rows.append({
                "external_name": ext,
                "canonical_name": "",
                "match_score": 0.0,
                "status": "unmatched",
            })

    write_csv(
        CROSSWALK,
        ["external_name", "canonical_name", "match_score", "status"],
        crosswalk_rows,
    )
    matched = sum(1 for r in crosswalk_rows if r["status"] in ("exact", "auto"))
    review  = sum(1 for r in crosswalk_rows if r["status"] == "review")
    unmatched = sum(1 for r in crosswalk_rows if r["status"] == "unmatched")
    print(f"  crosswalk.csv: {len(crosswalk_rows)} names — "
          f"{matched} auto ✓  {review} review ⚠  {unmatched} unmatched ✗")
    return mapping


# ══════════════════════════════════════════════════════════════════════════
# FASE 1 — regions, manufacturers, distributors
# ══════════════════════════════════════════════════════════════════════════

def gen_regions() -> int:
    h, rows = load_wb_rows(STORES_XLS, "STORES")
    state_i = next(i for i, x in enumerate(h) if "state" in x.lower())
    status_i = next(i for i, x in enumerate(h) if "status" in x.lower())

    seen: set[str] = set()
    out: list[dict] = []
    for r in rows:
        state_raw = clean_str(r[state_i])
        status    = clean_str(r[status_i])
        if not state_raw or status != "OPEN":
            continue
        code = norm_state(state_raw)
        if code in seen:
            continue
        seen.add(code)
        out.append({
            "name":         STATE_TO_REGION.get(state_raw, STATE_TO_REGION.get(code, code)),
            "code":         code,
            "country_code": "US",
            "is_active":    "true",
        })

    return write_csv(
        TMPL / "regions.csv",
        ["name", "code", "country_code", "is_active"],
        out,
    )


def gen_manufacturers() -> int:
    seen: set[str] = set()
    out: list[dict] = []

    def add(raw_key: str):
        info = MANUFACTURER_MAP.get(raw_key.lower())
        if info and info["name"] not in seen:
            seen.add(info["name"])
            out.append({
                "name":         info["name"],
                "country_code": info["country_code"],
                "tax_id":       "",
                "website":      "",
                "is_active":    "true",
            })

    # Bridor
    add("bridor")

    # Bindi + Kimbo (from Revised DB Group=Coffee)
    try:
        h_b, rows_b = load_wb_rows(BINDI, "Data")
        grp_i = next(i for i, x in enumerate(h_b) if "group" in x.lower())
        for r in rows_b:
            g = clean_str(r[grp_i])
            if g and g.lower() not in BINDI_DISTRIBUTORS:
                add(g.lower())
    except Exception:
        pass

    try:
        h_r, rows_r = load_wb_rows(BINDI, "Revised Prod Database")
        grp_i = next(i for i, x in enumerate(h_r) if "group" in x.lower())
        for r in rows_r:
            g = clean_str(r[grp_i])
            if g and g.lower() == "coffee":
                add("coffee")
                break
    except Exception:
        pass

    # PreGel
    add("pregel")
    # Golden Waffles
    add("golden waffles")

    return write_csv(
        TMPL / "manufacturers.csv",
        ["name", "country_code", "tax_id", "website", "is_active"],
        out,
    )


def gen_distributors() -> int:
    seen: set[str] = set()
    out: list[dict] = []

    def add(raw: str):
        norm = DISTRIBUTOR_NORM.get(raw.lower())
        if norm and norm not in seen:
            seen.add(norm)
            out.append({
                "name":          norm,
                "country_code":  "US",
                "tax_id":        "",
                "contact_email": "",
                "contact_phone": "",
                "is_active":     "true",
            })

    # From Bridor
    try:
        h, rows = load_wb_rows(BRIDOR, "Data")
        dist_i = next(i for i, x in enumerate(h) if "distributor" in x.lower())
        for r in rows:
            d = clean_str(r[dist_i])
            if d:
                add(d)
    except Exception:
        pass

    # From Bindi Group (distributors only)
    try:
        h_b, rows_b = load_wb_rows(BINDI, "Data")
        grp_i = next(i for i, x in enumerate(h_b) if "group" in x.lower())
        for r in rows_b:
            g = clean_str(r[grp_i])
            if g and g.lower() in BINDI_DISTRIBUTORS:
                add(g)
    except Exception:
        pass

    return write_csv(
        TMPL / "distributors.csv",
        ["name", "country_code", "tax_id", "contact_email", "contact_phone", "is_active"],
        out,
    )


# ══════════════════════════════════════════════════════════════════════════
# FASE 2 — stores_regions, ingredient_prices, supply_routes
# ══════════════════════════════════════════════════════════════════════════

def gen_stores_regions() -> int:
    h, rows = load_wb_rows(STORES_XLS, "STORES")
    new_code_i = next(i for i, x in enumerate(h) if "new" in x.lower() and "code" in x.lower())
    state_i    = next(i for i, x in enumerate(h) if x.lower() == "state")
    status_i   = next(i for i, x in enumerate(h) if "status" in x.lower())

    out: list[dict] = []
    for r in rows:
        code   = clean_str(r[new_code_i])
        state  = clean_str(r[state_i])
        status = clean_str(r[status_i])
        if not code or not state or status != "OPEN":
            continue
        out.append({
            "store_code":  code,
            "region_code": norm_state(state),
        })

    return write_csv(TMPL / "stores_regions.csv", ["store_code", "region_code"], out)


def _bindi_prices(crosswalk: dict) -> list[dict]:
    """Extract ingredient prices from Bindi Products Database."""
    rows_out: list[dict] = []
    try:
        h, rows = load_wb_rows(BINDI, "Products Database")
        item_i  = next(i for i, x in enumerate(h) if "item" in x.lower())
        grp_i   = next(i for i, x in enumerate(h) if "group" in x.lower())
        ppu_i   = next(i for i, x in enumerate(h) if "price per unit" in x.lower())
        ppc_i   = next(i for i, x in enumerate(h) if "price per case" in x.lower())
        dsd_i   = next((i for i, x in enumerate(h) if "dsd" in x.lower()), None)
        date_i  = next((i for i, x in enumerate(h) if x.lower() == "date"), None)
        bpc_i   = next(i for i, x in enumerate(h) if "bags" in x.lower() or "units per case" in x.lower())
    except (StopIteration, ValueError):
        return []

    # Aggregate: {ext_name_lower: {canonical, prices[], dates[]}}
    agg: dict[str, dict] = {}
    for r in rows:
        item = clean_str(r[item_i])
        if not item:
            continue
        ext_clean = clean_external_name(item)
        canonical = crosswalk.get(ext_clean.lower())
        if not canonical:
            continue

        # Prefer DSD price as qargo_price if cheaper
        ppc   = clean_price(r[ppc_i])
        dsd   = clean_price(r[dsd_i]) if dsd_i is not None else None
        ppu   = clean_price(r[ppu_i])
        bpc   = clean_price(r[bpc_i])

        # Best unit price
        if ppu and ppu > 0:
            unit_price = ppu
        elif ppc and bpc and bpc > 0:
            unit_price = round(ppc / bpc, 4)
        else:
            continue

        raw_date = r[date_i] if date_i is not None else None
        dt = raw_date.date() if isinstance(raw_date, datetime) else (
            raw_date if isinstance(raw_date, date) else None
        )

        key = canonical.lower()
        if key not in agg:
            agg[key] = {"canonical": canonical, "prices": [], "dates": []}
        agg[key]["prices"].append(unit_price)
        if dt:
            agg[key]["dates"].append(dt)

    for info in agg.values():
        eff_date = max(info["dates"]).isoformat() if info["dates"] else today_iso()
        rows_out.append({
            "ingredient_name": info["canonical"],
            "purchase_price":  round(statistics.median(info["prices"]), 4),
            "currency_code":   "USD",
            "source":          "bindi_2026",
            "effective_date":  eff_date,
        })
    return rows_out


def _pregel_prices(crosswalk: dict) -> list[dict]:
    rows_out: list[dict] = []
    try:
        h, rows = load_wb_rows(PREGEL, "Products Database")
        item_i = next(i for i, x in enumerate(h) if "item" in x.lower())
        ppu_i  = next(i for i, x in enumerate(h) if "price per unit" in x.lower())
        ppc_i  = next(i for i, x in enumerate(h) if "price per case" in x.lower())
        bpc_i  = next(i for i, x in enumerate(h) if "bags" in x.lower() or "units per case" in x.lower())
        date_i = next((i for i, x in enumerate(h) if x.lower() == "date"), None)
    except (StopIteration, ValueError):
        return []

    agg: dict[str, dict] = {}
    for r in rows:
        item = clean_str(r[item_i])
        if not item:
            continue
        canonical = crosswalk.get(clean_external_name(item).lower())
        if not canonical:
            continue
        ppu = clean_price(r[ppu_i])
        ppc = clean_price(r[ppc_i])
        bpc = clean_price(r[bpc_i])
        unit_price = ppu or (round(ppc / bpc, 4) if ppc and bpc and bpc > 0 else None)
        if not unit_price:
            continue
        raw_date = r[date_i] if date_i is not None else None
        dt = raw_date.date() if isinstance(raw_date, datetime) else (
            raw_date if isinstance(raw_date, date) else None
        )
        key = canonical.lower()
        if key not in agg:
            agg[key] = {"canonical": canonical, "prices": [], "dates": []}
        agg[key]["prices"].append(unit_price)
        if dt:
            agg[key]["dates"].append(dt)

    for info in agg.values():
        rows_out.append({
            "ingredient_name": info["canonical"],
            "purchase_price":  round(statistics.median(info["prices"]), 4),
            "currency_code":   "USD",
            "source":          "pregel_2026",
            "effective_date":  max(info["dates"]).isoformat() if info["dates"] else today_iso(),
        })
    return rows_out


def _waffles_prices(crosswalk: dict) -> list[dict]:
    rows_out: list[dict] = []
    try:
        h, rows = load_wb_rows(WAFFLES, "Products Database")
        item_i = next(i for i, x in enumerate(h) if "item" in x.lower())
        ppu_i  = next(i for i, x in enumerate(h) if "price per unit" in x.lower())
        ppc_i  = next(i for i, x in enumerate(h) if "price per case" in x.lower())
        bpc_i  = next(i for i, x in enumerate(h) if "bags" in x.lower() or "units per case" in x.lower())
        date_i = next((i for i, x in enumerate(h) if x.lower() == "date"), None)
    except (StopIteration, ValueError):
        return []

    agg: dict[str, dict] = {}
    for r in rows:
        item = clean_str(r[item_i])
        if not item:
            continue
        canonical = crosswalk.get(clean_external_name(item).lower())
        if not canonical:
            continue
        ppu = clean_price(r[ppu_i])
        ppc = clean_price(r[ppc_i])
        bpc = clean_price(r[bpc_i])
        unit_price = ppu or (round(ppc / bpc, 4) if ppc and bpc and bpc > 0 else None)
        if not unit_price:
            continue
        raw_date = r[date_i] if date_i is not None else None
        dt = raw_date.date() if isinstance(raw_date, datetime) else (
            raw_date if isinstance(raw_date, date) else None
        )
        key = canonical.lower()
        if key not in agg:
            agg[key] = {"canonical": canonical, "prices": [], "dates": []}
        agg[key]["prices"].append(unit_price)
        if dt:
            agg[key]["dates"].append(dt)

    for info in agg.values():
        rows_out.append({
            "ingredient_name": info["canonical"],
            "purchase_price":  round(statistics.median(info["prices"]), 4),
            "currency_code":   "USD",
            "source":          "golden_waffles_2026",
            "effective_date":  max(info["dates"]).isoformat() if info["dates"] else today_iso(),
        })
    return rows_out


def _bridor_prices(crosswalk: dict) -> list[dict]:
    rows_out: list[dict] = []
    try:
        h, rows = load_wb_rows(BRIDOR, "Data")
        desc_i  = next(i for i, x in enumerate(h) if "description" in x.lower())
        price_i = next(i for i, x in enumerate(h) if "total price" in x.lower())
        cases_i = next(i for i, x in enumerate(h) if "cases ordered" in x.lower())
    except (StopIteration, ValueError):
        return []

    agg: dict[str, dict] = {}
    for r in rows:
        desc = clean_str(r[desc_i])
        if not desc:
            continue
        canonical = crosswalk.get(clean_external_name(desc).lower())
        if not canonical:
            continue
        total = clean_price(r[price_i])
        cases = clean_price(r[cases_i])
        if not total or not cases or cases == 0:
            continue
        unit_price = round(total / cases, 4)
        key = canonical.lower()
        if key not in agg:
            agg[key] = {"canonical": canonical, "prices": []}
        agg[key]["prices"].append(unit_price)

    for info in agg.values():
        rows_out.append({
            "ingredient_name": info["canonical"],
            "purchase_price":  round(statistics.median(info["prices"]), 4),
            "currency_code":   "USD",
            "source":          "bridor_q4_2025",
            "effective_date":  "2025-10-01",
        })
    return rows_out


def _raw_ingredients_prices() -> list[dict]:
    """165 canonical ingredient prices from data/raw/ingredients.xlsx.
    Names are already canonical — no crosswalk needed.
    These form the BASE price layer; external sources override where available."""
    rows_out: list[dict] = []
    if not RAW_INGREDIENTS.exists():
        return rows_out
    wb    = openpyxl.load_workbook(RAW_INGREDIENTS, data_only=True)
    ws    = wb.active
    h     = [c.value for c in ws[1]]
    ni    = next(i for i, x in enumerate(h) if x and x.lower() == "name")
    ppi   = next(i for i, x in enumerate(h) if x and "purchase_price" in x.lower())
    for r in ws.iter_rows(min_row=2, values_only=True):
        name  = clean_str(r[ni])
        price = clean_price(r[ppi])
        if not name or not price:
            continue
        rows_out.append({
            "ingredient_name": name,
            "purchase_price":  price,
            "currency_code":   "USD",
            "source":          "raw_ingredients_v1",
            "effective_date":  today_iso(),
        })
    return rows_out


def gen_ingredient_prices(crosswalk: dict) -> int:
    # Base layer: all 165 canonical ingredients from raw/
    base_rows = _raw_ingredients_prices()
    # Override layer: supplier-specific prices from files-to-explore (more specific)
    override_rows = (
        _bindi_prices(crosswalk)
        + _pregel_prices(crosswalk)
        + _waffles_prices(crosswalk)
        + _bridor_prices(crosswalk)
    )

    # Merge: start with base, override with supplier prices
    merged: dict[str, dict] = {}
    for row in base_rows:
        merged[row["ingredient_name"].lower()] = row
    for row in override_rows:
        key = row["ingredient_name"].lower()
        # Supplier price overrides base only if cheaper (better deal for Qargo)
        if key not in merged or row["purchase_price"] < merged[key]["purchase_price"]:
            merged[key] = row

    return write_csv(
        TMPL / "ingredient_prices.csv",
        ["ingredient_name", "purchase_price", "currency_code", "source", "effective_date"],
        list(merged.values()),
    )


def gen_supply_routes(crosswalk: dict) -> int:
    seen: set[tuple] = set()
    out: list[dict] = []

    def add(ingredient: str, manufacturer: str, distributor: str | None, is_direct: bool):
        key = (ingredient.lower(), manufacturer.lower(), (distributor or "").lower())
        if key in seen:
            return
        seen.add(key)
        out.append({
            "ingredient_name":   ingredient,
            "manufacturer_name": manufacturer,
            "distributor_name":  distributor or "",
            "is_direct":         "true" if is_direct else "false",
            "is_active":         "true",
        })

    # Bindi
    try:
        h_b, rows_b = load_wb_rows(BINDI, "Products Database")
        item_i = next(i for i, x in enumerate(h_b) if "item" in x.lower())
        grp_i  = next(i for i, x in enumerate(h_b) if "group" in x.lower())
        ppc_i  = next(i for i, x in enumerate(h_b) if "price per case" in x.lower())
        dsd_i  = next((i for i, x in enumerate(h_b) if "dsd" in x.lower()), None)
        for r in rows_b:
            item = clean_str(r[item_i])
            grp  = clean_str(r[grp_i])
            if not item or not grp:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if not canonical:
                continue
            grp_lower = grp.lower()
            if grp_lower not in BINDI_DISTRIBUTORS:
                # Bindi direct
                ppc = clean_price(r[ppc_i])
                dsd = clean_price(r[dsd_i]) if dsd_i is not None else None
                is_direct = dsd is not None and ppc is not None and abs(dsd - ppc) < 0.01
                add(canonical, "Bindi", None if is_direct else grp, is_direct)
            else:
                dist_norm = DISTRIBUTOR_NORM.get(grp_lower, grp.title())
                add(canonical, "Bindi", dist_norm, False)
    except Exception:
        pass

    # Bridor
    try:
        h_br, rows_br = load_wb_rows(BRIDOR, "Data")
        desc_i = next(i for i, x in enumerate(h_br) if "description" in x.lower())
        dist_i = next(i for i, x in enumerate(h_br) if "distributor" in x.lower())
        for r in rows_br:
            desc = clean_str(r[desc_i])
            dist = clean_str(r[dist_i])
            if not desc:
                continue
            canonical = crosswalk.get(clean_external_name(desc).lower())
            if not canonical:
                continue
            dist_norm = DISTRIBUTOR_NORM.get((dist or "").lower(), (dist or "").title())
            add(canonical, "Bridor", dist_norm or None, False)
    except Exception:
        pass

    # PreGel (distributor unknown — mark empty)
    try:
        h_p, rows_p = load_wb_rows(PREGEL, "Products Database")
        item_i = next(i for i, x in enumerate(h_p) if "item" in x.lower())
        for r in rows_p:
            item = clean_str(r[item_i])
            if not item:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if canonical:
                add(canonical, "PreGel", None, False)
    except Exception:
        pass

    # Golden Waffles (distributor unknown)
    try:
        h_g, rows_g = load_wb_rows(WAFFLES, "Products Database")
        item_i = next(i for i, x in enumerate(h_g) if "item" in x.lower())
        for r in rows_g:
            item = clean_str(r[item_i])
            if not item:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if canonical:
                add(canonical, "Golden Waffles", None, False)
    except Exception:
        pass

    return write_csv(
        TMPL / "supply_routes.csv",
        ["ingredient_name", "manufacturer_name", "distributor_name", "is_direct", "is_active"],
        out,
    )


# ══════════════════════════════════════════════════════════════════════════
# FASE 3 — ingredient_supplier_refs, supplier_unit_conversions, route_prices
# ══════════════════════════════════════════════════════════════════════════

def gen_ingredient_supplier_refs(crosswalk: dict) -> int:
    seen: set[tuple] = set()
    out: list[dict] = []

    def add(canonical, manufacturer, distributor, ext_name, ext_code, purchase_unit, units_per_pack):
        key = (canonical.lower(), manufacturer.lower(), (distributor or "").lower(),
               (ext_code or "").lower())
        if key in seen:
            return
        seen.add(key)
        out.append({
            "ingredient_name":   canonical,
            "manufacturer_name": manufacturer,
            "distributor_name":  distributor or "",
            "external_name":     ext_name,
            "external_code":     ext_code or "",
            "purchase_unit":     purchase_unit or "",
            "units_per_pack":    units_per_pack or "",
        })

    def infer_pu(bpc, kg, lb, gal) -> str:
        bpc_int = int(bpc) if bpc and bpc == int(bpc) else bpc
        if gal and gal > 0:
            return f"Case {bpc_int} × {round(gal/bpc,2)}gal" if bpc else f"{gal}gal"
        if kg and kg > 0:
            return f"Case {bpc_int} × {round(kg/bpc,2)}kg" if bpc else f"{kg}kg"
        if lb and lb > 0:
            return f"Case {bpc_int} × {round(lb/bpc,2)}lb" if bpc else f"{lb}lb"
        return f"Case {bpc_int} units" if bpc else "unit"

    # Bindi Products Database
    try:
        h, rows = load_wb_rows(BINDI, "Products Database")
        code_i = next(i for i, x in enumerate(h) if x.lower() == "code")
        item_i = next(i for i, x in enumerate(h) if "item" in x.lower())
        grp_i  = next(i for i, x in enumerate(h) if "group" in x.lower())
        bpc_i  = next(i for i, x in enumerate(h) if "bags" in x.lower() or "units per case" in x.lower())
        for r in rows:
            item = clean_str(r[item_i])
            grp  = clean_str(r[grp_i])
            if not item or not grp:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if not canonical:
                continue
            is_dist = grp.lower() in BINDI_DISTRIBUTORS
            manufacturer = "Bindi"
            distributor  = DISTRIBUTOR_NORM.get(grp.lower()) if is_dist else None
            bpc = clean_price(r[bpc_i])
            add(canonical, manufacturer, distributor,
                clean_external_name(item), clean_code(r[code_i]),
                infer_pu(bpc, None, None, None), bpc)
    except Exception:
        pass

    # Bindi Revised Prod Database (has weight conversions)
    try:
        h, rows = load_wb_rows(BINDI, "Revised Prod Database")
        code_i = next(i for i, x in enumerate(h) if x.lower() == "code")
        item_i = next(i for i, x in enumerate(h) if "item" in x.lower())
        grp_i  = next(i for i, x in enumerate(h) if "group" in x.lower())
        bpc_i  = next(i for i, x in enumerate(h) if "bags" in x.lower() or "units per case" in x.lower())
        kg_i   = next((i for i, x in enumerate(h) if "kg" in x.lower()), None)
        lb_i   = next((i for i, x in enumerate(h) if "pound" in x.lower() or " lb" in x.lower()), None)
        gal_i  = next((i for i, x in enumerate(h) if "gallon" in x.lower()), None)
        for r in rows:
            item = clean_str(r[item_i])
            grp  = clean_str(r[grp_i])
            if not item or not grp:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if not canonical:
                continue
            manufacturer = "Kimbo" if grp.lower() == "coffee" else "Bindi"
            distributor  = None
            bpc = clean_price(r[bpc_i])
            kg  = clean_price(r[kg_i])  if kg_i  else None
            lb  = clean_price(r[lb_i])  if lb_i  else None
            gal = clean_price(r[gal_i]) if gal_i else None
            pu  = infer_pu(bpc, kg, lb, gal) if bpc else ""
            add(canonical, manufacturer, distributor,
                clean_external_name(item), clean_code(r[code_i]), pu, bpc)
    except Exception:
        pass

    # PreGel
    try:
        h, rows = load_wb_rows(PREGEL, "Products Database")
        code_i = next(i for i, x in enumerate(h) if x.lower() == "code")
        item_i = next(i for i, x in enumerate(h) if "item" in x.lower())
        bpc_i  = next(i for i, x in enumerate(h) if "bags" in x.lower() or "units per case" in x.lower())
        for r in rows:
            item = clean_str(r[item_i])
            if not item:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if not canonical:
                continue
            bpc = clean_price(r[bpc_i])
            add(canonical, "PreGel", None,
                clean_external_name(item), clean_code(r[code_i]),
                infer_pu(bpc, None, None, None), bpc)
    except Exception:
        pass

    # Golden Waffles
    try:
        h, rows = load_wb_rows(WAFFLES, "Products Database")
        code_i = next(i for i, x in enumerate(h) if x.lower() == "code")
        item_i = next(i for i, x in enumerate(h) if "item" in x.lower())
        bpc_i  = next(i for i, x in enumerate(h) if "bags" in x.lower() or "units per case" in x.lower())
        for r in rows:
            item = clean_str(r[item_i])
            if not item:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if not canonical:
                continue
            bpc = clean_price(r[bpc_i])
            add(canonical, "Golden Waffles", None,
                clean_external_name(item), clean_code(r[code_i]),
                infer_pu(bpc, None, None, None), bpc)
    except Exception:
        pass

    # Bridor
    try:
        h, rows = load_wb_rows(BRIDOR, "Data")
        num_i  = next(i for i, x in enumerate(h) if "item number" in x.lower())
        desc_i = next(i for i, x in enumerate(h) if "description" in x.lower())
        dist_i = next(i for i, x in enumerate(h) if "distributor" in x.lower())
        for r in rows:
            desc = clean_str(r[desc_i])
            dist = clean_str(r[dist_i])
            if not desc:
                continue
            canonical = crosswalk.get(clean_external_name(desc).lower())
            if not canonical:
                continue
            dist_norm = DISTRIBUTOR_NORM.get((dist or "").lower(), (dist or "").title()) or None
            add(canonical, "Bridor", dist_norm,
                clean_external_name(desc), clean_code(r[num_i]), "Case", None)
    except Exception:
        pass

    return write_csv(
        TMPL / "ingredient_supplier_refs.csv",
        ["ingredient_name","manufacturer_name","distributor_name",
         "external_name","external_code","purchase_unit","units_per_pack"],
        out,
    )


def _raw_unit_conversions() -> list[dict]:
    """Pull from data/raw/ingredients.xlsx → purchase_unit / conversion_factor.
    Represents: 1 purchase_unit = conversion_factor usage_units.
    Also pulls data/raw/conversions.xlsx (recipe-unit conversions).
    No manufacturer/distributor — these are catalog-level conversions."""
    rows_out: list[dict] = []

    # From ingredients.xlsx: purchase_unit → usage_unit conversion
    if RAW_INGREDIENTS.exists():
        wb  = openpyxl.load_workbook(RAW_INGREDIENTS, data_only=True)
        ws  = wb.active
        h   = [c.value for c in ws[1]]
        ni  = next(i for i, x in enumerate(h) if x and x.lower() == "name")
        pui = next(i for i, x in enumerate(h) if x and "purchase_unit" in x.lower())
        uui = next(i for i, x in enumerate(h) if x and "usage_unit" in x.lower())
        cfi = next(i for i, x in enumerate(h) if x and "conversion_factor" in x.lower())
        for r in ws.iter_rows(min_row=2, values_only=True):
            name = clean_str(r[ni])
            pu   = clean_str(r[pui])
            uu   = clean_str(r[uui])
            cf   = clean_price(r[cfi])
            if not name or not uu or not cf or cf <= 0:
                continue
            rows_out.append({
                "ingredient_name":   name,
                "manufacturer_name": "",
                "distributor_name":  "",
                "recipe_unit":       uu,
                "purchase_qty":      1,
                "recipe_qty":        cf,
                "notes":             f"1 {pu or 'pack'} = {cf} {uu} (from raw/ingredients.xlsx)",
            })

    # From conversions.xlsx: recipe_unit → ml_or_g conversion
    if RAW_CONVERSIONS.exists():
        wb2  = openpyxl.load_workbook(RAW_CONVERSIONS, data_only=True)
        ws2  = wb2.active
        h2   = [c.value for c in ws2[1]]
        ini  = next(i for i, x in enumerate(h2) if x and "ingredient" in x.lower())
        rui  = next(i for i, x in enumerate(h2) if x and "recipe_unit" in x.lower())
        eqi  = next(i for i, x in enumerate(h2) if x and "equivalent" in x.lower())
        noti = next((i for i, x in enumerate(h2) if x and "note" in x.lower()), None)
        for r in ws2.iter_rows(min_row=2, values_only=True):
            name = clean_str(r[ini])
            ru   = clean_str(r[rui])
            eq   = clean_price(r[eqi])
            note = clean_str(r[noti]) if noti is not None else ""
            if not name or not ru or not eq or eq <= 0:
                continue
            rows_out.append({
                "ingredient_name":   name,
                "manufacturer_name": "",
                "distributor_name":  "",
                "recipe_unit":       ru,
                "purchase_qty":      1,
                "recipe_qty":        eq,
                "notes":             note or f"1 {ru} = {eq} usage_units (from raw/conversions.xlsx)",
            })

    return rows_out


def gen_supplier_unit_conversions(crosswalk: dict) -> int:
    seen: set[tuple] = set()
    out: list[dict] = []

    def add(canonical, manufacturer, distributor, recipe_unit, purchase_qty, recipe_qty, note=""):
        if not recipe_qty or recipe_qty <= 0 or not purchase_qty or purchase_qty <= 0:
            return
        key = (canonical.lower(), manufacturer.lower(), (distributor or "").lower(), recipe_unit)
        if key in seen:
            return
        seen.add(key)
        out.append({
            "ingredient_name":   canonical,
            "manufacturer_name": manufacturer,
            "distributor_name":  distributor or "",
            "recipe_unit":       recipe_unit,
            "purchase_qty":      round(float(purchase_qty), 4),
            "recipe_qty":        round(float(recipe_qty), 4),
            "notes":             note,
        })

    # Bindi Revised Prod Database — has weight columns
    try:
        h, rows = load_wb_rows(BINDI, "Revised Prod Database")
        item_i = next(i for i, x in enumerate(h) if "item" in x.lower())
        grp_i  = next(i for i, x in enumerate(h) if "group" in x.lower())
        bpc_i  = next(i for i, x in enumerate(h) if "bags" in x.lower() or "units per case" in x.lower())
        kg_i   = next((i for i, x in enumerate(h) if "kg" in x.lower()), None)
        lb_i   = next((i for i, x in enumerate(h) if "pound" in x.lower() or " lb" in x.lower()), None)
        gal_i  = next((i for i, x in enumerate(h) if "gallon" in x.lower()), None)

        for r in rows:
            item = clean_str(r[item_i])
            grp  = clean_str(r[grp_i])
            if not item or not grp:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if not canonical:
                continue
            manufacturer = "Kimbo" if grp.lower() == "coffee" else "Bindi"
            bpc = clean_price(r[bpc_i])
            kg  = clean_price(r[kg_i])  if kg_i  else None
            lb  = clean_price(r[lb_i])  if lb_i  else None
            gal = clean_price(r[gal_i]) if gal_i else None

            if gal and gal > 0:
                add(canonical, manufacturer, None, "ml", bpc, gal * 3785.41,
                    f"{bpc} units/case = {gal}gal")
            elif kg and kg > 0:
                add(canonical, manufacturer, None, "g", bpc, kg * 1000,
                    f"{bpc} units/case = {kg}kg")
            elif lb and lb > 0:
                add(canonical, manufacturer, None, "g", bpc, lb * 453.592,
                    f"{bpc} units/case = {lb}lb")
            elif bpc and bpc > 0:
                add(canonical, manufacturer, None, "unit", bpc, bpc,
                    "unit only, no weight data")
    except Exception:
        pass

    # PreGel — units only
    try:
        h, rows = load_wb_rows(PREGEL, "Products Database")
        item_i = next(i for i, x in enumerate(h) if "item" in x.lower())
        bpc_i  = next(i for i, x in enumerate(h) if "bags" in x.lower() or "units per case" in x.lower())
        for r in rows:
            item = clean_str(r[item_i])
            if not item:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if not canonical:
                continue
            bpc = clean_price(r[bpc_i])
            if bpc and bpc > 0:
                add(canonical, "PreGel", None, "unit", bpc, bpc, "unit only")
    except Exception:
        pass

    # Golden Waffles — units only
    try:
        h, rows = load_wb_rows(WAFFLES, "Products Database")
        item_i = next(i for i, x in enumerate(h) if "item" in x.lower())
        bpc_i  = next(i for i, x in enumerate(h) if "bags" in x.lower() or "units per case" in x.lower())
        for r in rows:
            item = clean_str(r[item_i])
            if not item:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if not canonical:
                continue
            bpc = clean_price(r[bpc_i])
            if bpc and bpc > 0:
                add(canonical, "Golden Waffles", None, "unit", bpc, bpc, "unit only")
    except Exception:
        pass

    # raw/ingredients.xlsx + raw/conversions.xlsx (catalog-level conversions)
    for row in _raw_unit_conversions():
        key = (row["ingredient_name"].lower(), "", "", row["recipe_unit"])
        if key not in seen:
            seen.add(key)
            out.append(row)

    return write_csv(
        TMPL / "supplier_unit_conversions.csv",
        ["ingredient_name","manufacturer_name","distributor_name",
         "recipe_unit","purchase_qty","recipe_qty","notes"],
        out,
    )


def gen_supply_route_prices(crosswalk: dict) -> int:
    seen: set[tuple] = set()
    out: list[dict] = []

    def add(canonical, manufacturer, distributor, list_price, qargo_price,
            source, valid_from):
        if not list_price or list_price <= 0:
            return
        qp = qargo_price if qargo_price and 0 < qargo_price <= list_price else list_price
        key = (canonical.lower(), manufacturer.lower(), (distributor or "").lower())
        # Allow update with lower qargo_price
        if key in seen:
            for row in out:
                if (row["ingredient_name"].lower() == canonical.lower() and
                        row["manufacturer_name"].lower() == manufacturer.lower() and
                        row["distributor_name"].lower() == (distributor or "").lower()):
                    if qp < float(row["qargo_price"]):
                        row["qargo_price"]  = round(qp, 4)
                        row["list_price"]   = round(list_price, 4)
                        row["valid_from"]   = valid_from
                        row["source"]       = source
            return
        seen.add(key)
        out.append({
            "ingredient_name":   canonical,
            "manufacturer_name": manufacturer,
            "distributor_name":  distributor or "",
            "list_price":        round(list_price, 4),
            "qargo_price":       round(qp, 4),
            "currency_code":     "USD",
            "price_unit":        "per case",
            "valid_from":        valid_from,
            "source":            source,
            "created_by":        "etl_migration_2026",
        })

    # Bindi
    try:
        h, rows = load_wb_rows(BINDI, "Products Database")
        item_i = next(i for i, x in enumerate(h) if "item" in x.lower())
        grp_i  = next(i for i, x in enumerate(h) if "group" in x.lower())
        ppc_i  = next(i for i, x in enumerate(h) if "price per case" in x.lower())
        dsd_i  = next((i for i, x in enumerate(h) if "dsd" in x.lower()), None)
        date_i = next((i for i, x in enumerate(h) if x.lower() == "date"), None)

        for r in rows:
            item = clean_str(r[item_i])
            grp  = clean_str(r[grp_i])
            if not item or not grp:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if not canonical:
                continue
            ppc = clean_price(r[ppc_i])
            dsd = clean_price(r[dsd_i]) if dsd_i is not None else None
            raw_date = r[date_i] if date_i is not None else None
            dt = raw_date.date() if isinstance(raw_date, datetime) else (
                raw_date if isinstance(raw_date, date) else None
            )
            valid_from = dt.isoformat() if dt else today_iso()
            is_dist = grp.lower() in BINDI_DISTRIBUTORS
            manufacturer = "Bindi"
            distributor  = DISTRIBUTOR_NORM.get(grp.lower()) if is_dist else None
            qargo = dsd if (dsd and ppc and dsd < ppc) else ppc
            add(canonical, manufacturer, distributor, ppc, qargo,
                "bindi_products_db_2026", valid_from)
    except Exception:
        pass

    # Bridor
    try:
        h, rows = load_wb_rows(BRIDOR, "Data")
        desc_i  = next(i for i, x in enumerate(h) if "description" in x.lower())
        price_i = next(i for i, x in enumerate(h) if "total price" in x.lower())
        cases_i = next(i for i, x in enumerate(h) if "cases ordered" in x.lower())
        dist_i  = next(i for i, x in enumerate(h) if "distributor" in x.lower())
        for r in rows:
            desc = clean_str(r[desc_i])
            if not desc:
                continue
            canonical = crosswalk.get(clean_external_name(desc).lower())
            if not canonical:
                continue
            total = clean_price(r[price_i])
            cases = clean_price(r[cases_i])
            if not total or not cases or cases == 0:
                continue
            dist = clean_str(r[dist_i])
            dist_norm = DISTRIBUTOR_NORM.get((dist or "").lower(), (dist or "").title()) or None
            list_price = round(total / cases, 4)
            add(canonical, "Bridor", dist_norm, list_price, list_price,
                "bridor_q4_2025", "2025-10-01")
    except Exception:
        pass

    # PreGel
    try:
        h, rows = load_wb_rows(PREGEL, "Products Database")
        item_i = next(i for i, x in enumerate(h) if "item" in x.lower())
        ppc_i  = next(i for i, x in enumerate(h) if "price per case" in x.lower())
        date_i = next((i for i, x in enumerate(h) if x.lower() == "date"), None)
        for r in rows:
            item = clean_str(r[item_i])
            if not item:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if not canonical:
                continue
            ppc = clean_price(r[ppc_i])
            raw_date = r[date_i] if date_i is not None else None
            dt = raw_date.date() if isinstance(raw_date, datetime) else (
                raw_date if isinstance(raw_date, date) else None
            )
            add(canonical, "PreGel", None, ppc, ppc, "pregel_2026",
                dt.isoformat() if dt else today_iso())
    except Exception:
        pass

    # Golden Waffles
    try:
        h, rows = load_wb_rows(WAFFLES, "Products Database")
        item_i = next(i for i, x in enumerate(h) if "item" in x.lower())
        ppc_i  = next(i for i, x in enumerate(h) if "price per case" in x.lower())
        date_i = next((i for i, x in enumerate(h) if x.lower() == "date"), None)
        for r in rows:
            item = clean_str(r[item_i])
            if not item:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if not canonical:
                continue
            ppc = clean_price(r[ppc_i])
            raw_date = r[date_i] if date_i is not None else None
            dt = raw_date.date() if isinstance(raw_date, datetime) else (
                raw_date if isinstance(raw_date, date) else None
            )
            add(canonical, "Golden Waffles", None, ppc, ppc,
                "golden_waffles_2026", dt.isoformat() if dt else today_iso())
    except Exception:
        pass

    return write_csv(
        TMPL / "supply_route_prices.csv",
        ["ingredient_name","manufacturer_name","distributor_name",
         "list_price","qargo_price","currency_code","price_unit",
         "valid_from","source","created_by"],
        out,
    )


# ══════════════════════════════════════════════════════════════════════════
# FASE 4 — supply_route_assignments (parcial)
# ══════════════════════════════════════════════════════════════════════════

def gen_supply_route_assignments(crosswalk: dict) -> int:
    seen: set[tuple] = set()
    out: list[dict] = []

    def add(store_code, canonical, manufacturer, distributor, valid_from):
        key = (store_code, canonical.lower(), manufacturer.lower(),
               (distributor or "").lower())
        if key in seen:
            return
        seen.add(key)
        out.append({
            "scope_type":       "store",
            "scope_code":       store_code,
            "ingredient_name":  canonical,
            "manufacturer_name":manufacturer,
            "distributor_name": distributor or "",
            "priority":         "1",
            "valid_from":       valid_from,
            "assigned_by":      "etl_migration_2026",
            "change_reason":    "initial_load",
            "notes":            "",
        })

    # Bindi Data — per-store purchasing records
    try:
        h, rows = load_wb_rows(BINDI, "Data")
        store_i = next(i for i, x in enumerate(h) if x.lower() == "store")
        item_i  = next(i for i, x in enumerate(h) if "item" in x.lower())
        grp_i   = next(i for i, x in enumerate(h) if "group" in x.lower())
        date_i  = next((i for i, x in enumerate(h) if x.lower() == "date"), None)

        for r in rows:
            store_raw = clean_str(r[store_i])
            item      = clean_str(r[item_i])
            grp       = clean_str(r[grp_i])
            if not store_raw or not item or not grp:
                continue
            if store_raw.lower() in SKIP_STORES:
                continue
            store_code = norm_store_name(store_raw)
            if not store_code:
                continue
            canonical = crosswalk.get(clean_external_name(item).lower())
            if not canonical:
                continue
            is_dist = grp.lower() in BINDI_DISTRIBUTORS
            manufacturer = "Bindi"
            distributor  = DISTRIBUTOR_NORM.get(grp.lower()) if is_dist else None
            raw_date = r[date_i] if date_i is not None else None
            dt = raw_date.date() if isinstance(raw_date, datetime) else (
                raw_date if isinstance(raw_date, date) else None
            )
            valid_from = dt.isoformat() if dt else today_iso()
            add(store_code, canonical, manufacturer, distributor, valid_from)
    except Exception:
        pass

    # Bridor Data — per-store (QC Store column)
    try:
        h, rows = load_wb_rows(BRIDOR, "Data")
        store_i = next(i for i, x in enumerate(h) if "store" in x.lower())
        desc_i  = next(i for i, x in enumerate(h) if "description" in x.lower())
        dist_i  = next(i for i, x in enumerate(h) if "distributor" in x.lower())
        for r in rows:
            store_raw = clean_str(r[store_i])
            desc      = clean_str(r[desc_i])
            dist      = clean_str(r[dist_i])
            if not store_raw or not desc:
                continue
            store_code = norm_store_name(store_raw)
            if not store_code:
                continue
            canonical = crosswalk.get(clean_external_name(desc).lower())
            if not canonical:
                continue
            dist_norm = DISTRIBUTOR_NORM.get((dist or "").lower(), (dist or "").title()) or None
            add(store_code, canonical, "Bridor", dist_norm, "2025-10-01")
    except Exception:
        pass

    return write_csv(
        TMPL / "supply_route_assignments.csv",
        ["scope_type","scope_code","ingredient_name","manufacturer_name",
         "distributor_name","priority","valid_from","assigned_by",
         "change_reason","notes"],
        out,
    )


# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description="ETL files-to-explore → templates")
    ap.add_argument("--phase", type=int, default=-1,
                    help="Run single phase (0-4). Omit for all phases.")
    args = ap.parse_args()
    run_all = args.phase == -1

    crosswalk: dict[str, str] = {}

    if args.phase == 0 or (run_all and not CROSSWALK.exists()):
        # Rebuild crosswalk only when explicitly requested or file missing
        print("\n── FASE 0: Crosswalk ──────────────────────────────────────")
        crosswalk = build_crosswalk()

    if run_all or args.phase >= 1:
        # Load crosswalk if not built in this run
        if not crosswalk and CROSSWALK.exists():
            with open(CROSSWALK, newline="", encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    if row["status"] in ("exact", "auto") and row["canonical_name"]:
                        crosswalk[row["external_name"].lower()] = row["canonical_name"]
            print(f"  Loaded {len(crosswalk)} crosswalk mappings from {CROSSWALK.name}")

    if run_all or args.phase == 1:
        print("\n── FASE 1: Entidades base ─────────────────────────────────")
        n = gen_regions();       print(f"  regions.csv              → {n} rows")
        n = gen_manufacturers(); print(f"  manufacturers.csv        → {n} rows")
        n = gen_distributors();  print(f"  distributors.csv         → {n} rows")

    if run_all or args.phase == 2:
        print("\n── FASE 2: Stores + precios + rutas ───────────────────────")
        n = gen_stores_regions();            print(f"  stores_regions.csv       → {n} rows")
        n = gen_ingredient_prices(crosswalk);print(f"  ingredient_prices.csv    → {n} rows")
        n = gen_supply_routes(crosswalk);    print(f"  supply_routes.csv        → {n} rows")

    if run_all or args.phase == 3:
        print("\n── FASE 3: Referencias + conversiones + precios de ruta ───")
        n = gen_ingredient_supplier_refs(crosswalk);  print(f"  ingredient_supplier_refs.csv     → {n} rows")
        n = gen_supplier_unit_conversions(crosswalk); print(f"  supplier_unit_conversions.csv    → {n} rows")
        n = gen_supply_route_prices(crosswalk);       print(f"  supply_route_prices.csv          → {n} rows")

    if run_all or args.phase == 4:
        print("\n── FASE 4: Asignaciones de ruta (parcial) ─────────────────")
        n = gen_supply_route_assignments(crosswalk);  print(f"  supply_route_assignments.csv     → {n} rows")

    print("\n✓ ETL completado. Archivos en data/templates/")
    print("  ⚠  Revisar crosswalk.csv: filas con status=review|unmatched")
    print("  supply_route_assignments.csv: priority=1 (todas rutas primarias)")


if __name__ == "__main__":
    main()
